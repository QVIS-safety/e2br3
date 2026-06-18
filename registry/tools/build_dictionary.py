#!/usr/bin/env python3
"""Build registry/dictionary/*.json from the spec source files in registry/sources/.

The generated JSON files are committed; rerun this script only when a source
spec document changes. Reading the MFDS xlsx requires openpyxl.
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
SOURCES_DIR = ROOT / "sources"
DICTIONARY_DIR = ROOT / "dictionary"

ICH_SOURCE = "ich-core-data-elements-v1.csv"
MFDS_SOURCE = "mfds-safety-r3-business-rules.xlsx"
FDA_SOURCE = "fda-core-regional-data-elements-v1.csv"
FDA_SEVERITY_SOURCE = "fda-rejection-warning-rules.csv"
XPATH_SOURCE = "mfds-icsr-element-xpath.csv"

CODE_PATTERN = re.compile(r"^[NCDEFGH](\.|$)")
NULL_FLAVOR_LABELS = ("NI", "MSK", "UNK", "NA", "ASKU", "NASK", "NINF", "PINF", "OTH")
ICH_CONFORMANCE_MAP = {
    "mandatory": "mandatory",
    "conditional-mandatory": "conditional_mandatory",
    "conditional mandatory": "conditional_mandatory",
    "optional": "optional",
    "required": "required",
}
MFDS_CONFORMANCE_MAP = {
    "필수": "mandatory",
    "조건부필수": "conditional_mandatory",
    "비필수": "optional",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace(" ", " ").strip()


def optional_value(value: Any) -> str | None:
    text = clean(value)
    return text if text and text != "-" else None


def set_optional(entry: dict[str, Any], key: str, value: Any) -> None:
    text = optional_value(value)
    if text is not None:
        entry[key] = text


def parse_ich_csv(text: str) -> list[dict[str, Any]]:
    rows = list(csv.reader(io.StringIO(text.lstrip("﻿"))))
    header_index = next(
        index for index, row in enumerate(rows) if "DATA ELEMENT NUMBER" in [clean(cell) for cell in row]
    )
    header = [clean(cell) for cell in rows[header_index]]
    column = {name: header.index(name) for name in header if name}
    null_flavor_columns = [
        (label, index) for index, label in enumerate(header) if label in NULL_FLAVOR_LABELS
    ]

    def cell(row: list[str], name: str) -> str:
        index = column[name]
        return clean(row[index]) if index < len(row) else ""

    entries: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[header_index + 1 :]:
        code = cell(row, "DATA ELEMENT NUMBER")
        name = cell(row, "DATA ELEMENT NAME")
        if code in ("", "-"):
            code = cell(row, "HEADER ELEMENT")
        if not CODE_PATTERN.match(code) or not name or code in seen:
            continue
        seen.add(code)

        conformance_raw = cell(row, "CONFORMANCE")
        normalized = ICH_CONFORMANCE_MAP.get(conformance_raw.lower().replace("–", "-"))
        entry: dict[str, Any] = {
            "code": code,
            "name": name,
            "section": code[0],
            "kind": "element" if normalized else "group",
        }
        if normalized:
            entry["conformance"] = normalized
        set_optional(entry, "data_type", cell(row, "DATA TYPE"))
        set_optional(entry, "max_length", cell(row, "MAX LENGTH"))
        set_optional(entry, "allowed_values", cell(row, "VALUE ALLOWED"))
        set_optional(entry, "oid", cell(row, "Code system OID"))
        null_flavors = [
            label
            for label, index in null_flavor_columns
            if index < len(row) and clean(row[index]).lower().startswith("yes")
        ]
        if null_flavors:
            entry["null_flavors"] = null_flavors
        entries.append(entry)
    return entries


def extract_ich_rules(text: str) -> dict[str, str]:
    """Pull the ICH business-rule prose per element code from the ICH table."""
    rows = list(csv.reader(io.StringIO(text.lstrip("﻿"))))
    header_index = next(
        index for index, row in enumerate(rows) if "DATA ELEMENT NUMBER" in [clean(cell) for cell in row]
    )
    header = [clean(cell) for cell in rows[header_index]]

    rules: dict[str, str] = {}
    for row in rows[header_index + 1 :]:
        def cell(name: str) -> str:
            index = header.index(name)
            return clean(row[index]) if index < len(row) else ""

        code = cell("DATA ELEMENT NUMBER")
        if code in ("", "-"):
            code = cell("HEADER ELEMENT")
        rule = optional_value(cell("ICH BUSINESS RULE"))
        if CODE_PATTERN.match(code) and rule and code not in rules:
            rules[code] = rule
    return rules


def parse_xpath_csv(text: str) -> dict[str, dict[str, str]]:
    """Map element code -> xpath / HL7 data type from the MFDS ICSR xPath table."""
    rows = list(csv.reader(io.StringIO(text.lstrip("﻿"))))

    def normalize(value: str) -> str:
        return "".join(clean(value).lower().split())

    header_index = next(
        index for index, row in enumerate(rows)
        if any(normalize(cell).startswith("elementnumber") for cell in row)
    )
    header = [normalize(cell) for cell in rows[header_index]]
    code_column = next(i for i, name in enumerate(header) if name.startswith("elementnumber"))
    xpath_column = next(i for i, name in enumerate(header) if name.startswith("xpath"))
    hl7_columns = [i for i, name in enumerate(header) if name.startswith("hl7datatype")]
    hl7_main, hl7_sub = hl7_columns[0], hl7_columns[1]

    mapping: dict[str, dict[str, str]] = {}
    for row in rows[header_index + 1 :]:
        def cell(index: int) -> str:
            return clean(row[index]) if index < len(row) else ""

        code = cell(code_column)
        if not CODE_PATTERN.match(code) or code in mapping:
            continue
        values = {}
        for key, index in (("xpath", xpath_column), ("hl7_data_type", hl7_main), ("hl7_component", hl7_sub)):
            value = optional_value(cell(index))
            if value:
                values[key] = value
        mapping[code] = values
    return mapping


def merge_xpath(entries: list[dict[str, Any]], mapping: dict[str, dict[str, str]]) -> int:
    annotated = 0
    for entry in entries:
        values = mapping.get(entry["code"])
        if values:
            entry.update(values)
            annotated += 1
    return annotated


FDA_CONFORMANCE_MAP = {
    "mandatory": "mandatory",
    "required": "mandatory",
    "conditional-mandatory": "conditional_mandatory",
    "conditional-required": "conditional_mandatory",
    "optional": "optional",
}
CONFORMANCE_STRICTNESS = ["mandatory", "conditional_mandatory", "optional"]


def read_fda_table(text: str) -> tuple[list[list[str]], list[str], dict[str, int]]:
    """Return (data rows, header, profile-name -> conformance column index)."""
    rows = list(csv.reader(io.StringIO(text.lstrip("﻿"))))
    header_index = next(
        index for index, row in enumerate(rows) if "DATA ELEMENT NUMBER" in [clean(cell) for cell in row]
    )
    header = [clean(cell) for cell in rows[header_index]]
    conformance_columns = [index for index, name in enumerate(header) if name == "CONFORMANCE"]
    profile_columns = dict(zip(("ich", "post_market", "pre_market", "vaers"), conformance_columns))
    return rows[header_index + 1 :], header, profile_columns


def fda_row_profiles(
    row: list[str], profile_columns: dict[str, int]
) -> dict[str, str]:
    profiles = {}
    for profile in ("post_market", "pre_market", "vaers"):
        index = profile_columns[profile]
        raw = clean(row[index]) if index < len(row) else ""
        normalized = FDA_CONFORMANCE_MAP.get(raw.lower().replace("–", "-"))
        if normalized:
            profiles[profile] = normalized
    return profiles


def merge_fda_profiles(ich_entries: list[dict[str, Any]], text: str) -> int:
    """Annotate ICH entries with FDA's per-route conformance from the FDA combined table.

    The FDA table restates every ICH element with Post-Market / Pre-Market /
    VAERS conformance columns; "-" means FDA states no deviation and adds no profile.
    """
    data_rows, header, profile_columns = read_fda_table(text)
    source_column = header.index("SOURCE")
    code_column = header.index("DATA ELEMENT NUMBER")

    rulings: dict[str, dict[str, str]] = {}
    for row in data_rows:
        if len(row) <= code_column or clean(row[source_column]).upper() != "ICH":
            continue
        code = clean(row[code_column])
        profiles = fda_row_profiles(row, profile_columns)
        if code and profiles:
            rulings.setdefault(code, profiles)

    annotated = 0
    for entry in ich_entries:
        profiles = rulings.get(entry["code"])
        if profiles and entry["kind"] == "element":
            entry.setdefault("profiles", {}).update(profiles)
            annotated += 1
    return annotated


def parse_fda_csv(text: str) -> list[dict[str, Any]]:
    """Build the FDA regional entries (SOURCE == FDA rows) from the FDA combined table."""
    data_rows, header, profile_columns = read_fda_table(text)
    null_flavor_columns = [
        (label, index) for index, label in enumerate(header) if label in NULL_FLAVOR_LABELS
    ]
    xpath_column = header.index("Value")

    def cell(row: list[str], index: int) -> str:
        return clean(row[index]) if index < len(row) else ""

    entries: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in data_rows:
        if cell(row, header.index("SOURCE")).upper() != "FDA":
            continue
        code = cell(row, header.index("DATA ELEMENT NUMBER"))
        name = cell(row, header.index("DATA ELEMENT NAME"))
        if not code.startswith("FDA.") or not name or code in seen:
            continue
        seen.add(code)

        profiles = fda_row_profiles(row, profile_columns)

        entry: dict[str, Any] = {
            "code": code,
            "name": name,
            "section": code.split(".")[1],
            "kind": "element" if profiles else "group",
        }
        if profiles:
            entry["conformance"] = min(profiles.values(), key=CONFORMANCE_STRICTNESS.index)
        set_optional(entry, "data_type", cell(row, header.index("DATA TYPE")))
        set_optional(entry, "max_length", cell(row, header.index("MAX LENGTH")))
        set_optional(entry, "allowed_values", cell(row, header.index("VALUES ALLOWED")))
        set_optional(entry, "oid", cell(row, header.index("Code system OID")))
        set_optional(entry, "hl7_data_type", cell(row, header.index("HL7 Data Type")))
        set_optional(entry, "hl7_component", cell(row, header.index("HL7 Data Type (sub component)")))
        if profiles:
            entry["profiles"] = profiles
        null_flavors = [
            label
            for label, index in null_flavor_columns
            if cell(row, index).lower().startswith("yes")
        ]
        if null_flavors:
            entry["null_flavors"] = null_flavors
        set_optional(entry, "xpath", cell(row, xpath_column))
        entries.append(entry)
    return entries


def extract_fda_rules(text: str) -> dict[str, str]:
    """Pull FDA's route-labelled rule prose per element code (ICH and FDA alike)."""
    data_rows, header, _ = read_fda_table(text)
    code_column = header.index("DATA ELEMENT NUMBER")
    rule_columns = {
        "Post-Market": header.index("Post-Market Business Rule"),
        "Pre-Market": header.index("Pre-Market Business Rule"),
        "VAERS": header.index("VAERS Business Rules"),
    }

    rules: dict[str, str] = {}
    for row in data_rows:
        code = clean(row[code_column]) if code_column < len(row) else ""
        if not (CODE_PATTERN.match(code) or code.startswith("FDA.")) or code in rules:
            continue
        labelled = [
            f"{label}: {clean(row[index])}"
            for label, index in rule_columns.items()
            if index < len(row) and optional_value(row[index])
        ]
        if labelled:
            rules[code] = "\n".join(labelled)
    return rules


def extract_fda_severity(text: str) -> dict[str, dict[str, str]]:
    """Map element code -> {severity, error_id, error_description} from the FDA
    rejection/warning rules table. Rows with neither mark are skipped."""
    rows = list(csv.reader(io.StringIO(text.lstrip("﻿"))))
    header_index = next(
        index for index, row in enumerate(rows)
        if "DATA ELEMENT NUMBER" in [clean(cell) for cell in row]
    )
    header = [clean(cell) for cell in rows[header_index]]
    col = {name: header.index(name) for name in header if name}
    rej_col = col["REJECTION, IF NOT MET"]
    warn_col = col["WARNING, IF NOT MET"]
    id_col = col.get("ERROR ID")
    desc_col = col.get("ERROR DESCRIPTION")

    mapping: dict[str, dict[str, str]] = {}
    for row in rows[header_index + 1 :]:
        def cell(index: int | None) -> str:
            return clean(row[index]) if index is not None and index < len(row) else ""

        code = cell(col["DATA ELEMENT NUMBER"])
        if not (CODE_PATTERN.match(code) or code.startswith("FDA.")) or code in mapping:
            continue
        rejection = bool(cell(rej_col))
        warning = bool(cell(warn_col))
        if not (rejection or warning):
            continue
        entry = {"severity": "rejection" if rejection else "warning"}
        if cell(id_col):
            entry["error_id"] = cell(id_col)
        if cell(desc_col):
            entry["error_description"] = cell(desc_col)
        mapping[code] = entry
    return mapping


def merge_fda_severity(entries: list[dict[str, Any]], mapping: dict[str, dict[str, str]]) -> int:
    annotated = 0
    for entry in entries:
        severity = mapping.get(entry["code"])
        if not severity or entry["kind"] != "element":
            continue
        entry["fda_severity"] = severity["severity"]
        if "error_id" in severity:
            entry["fda_error_id"] = severity["error_id"]
        annotated += 1
    return annotated


def normalize_header(value: Any) -> str:
    return clean(value).replace("\n", "").replace(" ", "")


def rows_as_dicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header = [normalize_header(cell) for cell in rows[0]]
    return [
        {header[index]: row[index] for index in range(min(len(header), len(row)))}
        for row in rows[1:]
    ]


def parse_mfds_sheets(
    sheet1_rows: list[list[Any]], sheet2_rows: list[list[Any]]
) -> list[dict[str, Any]]:
    """Build the official KR extension entries from the MFDS business-rule workbook.

    Sheet 2 ("신규 생성 항목") is the authoritative list of KR extension elements;
    sheet 1 ("전체 항목검증 룰") supplies their conformance and field metadata.
    """
    rules_by_code = {clean(row.get("ElementID")): row for row in rows_as_dicts(sheet1_rows)}

    entries: list[dict[str, Any]] = []
    for row in rows_as_dicts(sheet2_rows):
        code = clean(row.get("ElementID"))
        if not CODE_PATTERN.match(code):
            continue
        rule = rules_by_code.get(code, {})
        entry: dict[str, Any] = {
            "code": code,
            "name": clean(rule.get("항목명(영문)")) or clean(row.get("항목명(국문)")),
            "name_kr": clean(rule.get("항목명(국문)")) or clean(row.get("항목명(국문)")),
            "section": code[0],
            "kind": "element",
            "conformance": MFDS_CONFORMANCE_MAP.get(clean(rule.get("필수여부")), "optional"),
        }
        set_optional(entry, "data_type", rule.get("데이터유형"))
        set_optional(entry, "max_length", rule.get("최대길이"))
        set_optional(entry, "allowed_values", rule.get("허용치") or row.get("허용치"))
        set_optional(entry, "oid", rule.get("OID") or row.get("OID"))
        entries.append(entry)
    return entries


def extract_nullflavor_usage(sheet_rows: list[list[Any]]) -> dict[str, list[str]]:
    """Map element code -> allowed nullFlavor codes from the MFDS nullFlavor sheet."""
    mapping: dict[str, list[str]] = {}
    for row in rows_as_dicts(sheet_rows):
        code = clean(row.get("ElementID"))
        allowed = clean(row.get("허용치"))
        if not CODE_PATTERN.match(code) or not allowed:
            continue
        flavors = [token for token in re.split(r"[,\s]+", allowed) if token in NULL_FLAVOR_LABELS]
        if flavors:
            mapping[code] = flavors
    return mapping


def extract_vocabulary_usage(sheet_rows: list[list[Any]], vocabulary: str) -> dict[str, str]:
    """Map element code -> vocabulary name from an MFDS 'X 사용 항목' sheet."""
    mapping: dict[str, str] = {}
    for row in rows_as_dicts(sheet_rows):
        code = clean(row.get("ElementID"))
        if CODE_PATTERN.match(code):
            mapping[code] = vocabulary
    return mapping


def merge_nullflavors(entries: list[dict[str, Any]], mapping: dict[str, list[str]]) -> int:
    annotated = 0
    for entry in entries:
        flavors = mapping.get(entry["code"])
        if not flavors or entry["kind"] != "element":
            continue
        existing = entry.get("null_flavors", [])
        merged = existing + [f for f in flavors if f not in existing]
        entry["null_flavors"] = merged
        annotated += 1
    return annotated


def merge_vocabulary(entries: list[dict[str, Any]], mapping: dict[str, str]) -> int:
    annotated = 0
    for entry in entries:
        vocabulary = mapping.get(entry["code"])
        if vocabulary and entry["kind"] == "element":
            entry["vocabulary"] = vocabulary
            annotated += 1
    return annotated


def extract_mfds_rules(sheet1_rows: list[list[Any]]) -> dict[str, str]:
    """Pull the MFDS 항목검증룰 prose per element code (ICH and KR alike) from sheet 1."""
    rules: dict[str, str] = {}
    for row in rows_as_dicts(sheet1_rows):
        code = clean(row.get("ElementID"))
        rule = optional_value(row.get("항목검증룰"))
        if CODE_PATTERN.match(code) and rule and code not in rules:
            rules[code] = rule
    return rules


def merge_mfds_profiles(
    ich_entries: list[dict[str, Any]], sheet1_rows: list[list[Any]]
) -> int:
    """Annotate ICH entries with the MFDS profile conformance from workbook sheet 1.

    MFDS rules on ICH elements too (필수여부), not only on its KR extensions;
    this is recorded as profiles.mfds on the ICH dictionary entry.
    """
    rulings: dict[str, str] = {}
    for row in rows_as_dicts(sheet1_rows):
        code = clean(row.get("ElementID"))
        conformance = MFDS_CONFORMANCE_MAP.get(clean(row.get("필수여부")))
        if code and conformance:
            rulings[code] = conformance

    annotated = 0
    for entry in ich_entries:
        conformance = rulings.get(entry["code"])
        if conformance and entry["kind"] == "element":
            entry.setdefault("profiles", {})["mfds"] = conformance
            annotated += 1
    return annotated


def read_xlsx_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
    workbook.close()
    return rows


def write_dictionary(name: str, authority: str, source: str, entries: list[dict[str, Any]]) -> Path:
    DICTIONARY_DIR.mkdir(exist_ok=True)
    path = DICTIONARY_DIR / name
    payload = {"authority": authority, "source": f"sources/{source}", "entries": entries}
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return path


def write_rules(name: str, authority: str, source: str, rules: dict[str, str]) -> Path:
    rules_dir = DICTIONARY_DIR / "rules"
    rules_dir.mkdir(parents=True, exist_ok=True)
    path = rules_dir / name
    payload = {"authority": authority, "source": f"sources/{source}", "rules": rules}
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return path


VOCABULARY_SHEETS = {
    "4. MedDRA 사용 항목": "MedDRA",
    "5. WHODrug, 식약처코드 사용 항목": "WHODrug",
    "6-1. 국가코드 사용항목": "ISO3166",
    "6-2. 언어코드 사용 항목": "ISO639",
    "6-3. 성별코드 사용 항목": "sex",
    "6-4. UCUM 사용 항목": "UCUM",
    "6-5. EDQM 사용 항목": "EDQM",
}


def build_mfds_usage(mfds_path: Path) -> tuple[dict[str, list[str]], dict[str, str]]:
    """Load the MFDS nullFlavor + vocabulary usage sheets into merge-ready maps."""
    nullflavor = extract_nullflavor_usage(read_xlsx_rows(mfds_path, "3. nullflavor 사용 항목"))
    vocabulary: dict[str, str] = {}
    for sheet_name, vocab in VOCABULARY_SHEETS.items():
        vocabulary.update(extract_vocabulary_usage(read_xlsx_rows(mfds_path, sheet_name), vocab))
    return nullflavor, vocabulary


def main() -> int:
    mfds_path = SOURCES_DIR / MFDS_SOURCE
    mfds_sheet1 = read_xlsx_rows(mfds_path, "1. 전체 항목검증 룰")
    nullflavor_usage, vocabulary_usage = build_mfds_usage(mfds_path)

    ich_text = (SOURCES_DIR / ICH_SOURCE).read_text(encoding="utf-8")
    fda_text = (SOURCES_DIR / FDA_SOURCE).read_text(encoding="utf-8-sig")
    fda_severity = extract_fda_severity((SOURCES_DIR / FDA_SEVERITY_SOURCE).read_text(encoding="utf-8-sig"))
    xpath_map = parse_xpath_csv((SOURCES_DIR / XPATH_SOURCE).read_text(encoding="utf-8-sig"))

    ich_entries = parse_ich_csv(ich_text)
    mfds_annotated = merge_mfds_profiles(ich_entries, mfds_sheet1)
    fda_annotated = merge_fda_profiles(ich_entries, fda_text)
    xpath_annotated = merge_xpath(ich_entries, xpath_map)
    nf_annotated = merge_nullflavors(ich_entries, nullflavor_usage)
    vocab_annotated = merge_vocabulary(ich_entries, vocabulary_usage)
    sev_annotated = merge_fda_severity(ich_entries, fda_severity)
    path = write_dictionary("ich-e2br3.json", "ICH", ICH_SOURCE, ich_entries)
    elements = sum(1 for entry in ich_entries if entry["kind"] == "element")
    print(
        f"wrote {path} ({elements} elements, {len(ich_entries) - elements} groups,"
        f" {mfds_annotated} with an MFDS profile, {fda_annotated} with FDA profiles,"
        f" {xpath_annotated} with xpath, {nf_annotated} with MFDS nullFlavors,"
        f" {vocab_annotated} with a vocabulary, {sev_annotated} with FDA severity)"
    )

    mfds_entries = parse_mfds_sheets(
        mfds_sheet1,
        read_xlsx_rows(mfds_path, "2. 신규 생성 항목"),
    )
    xpath_annotated = merge_xpath(mfds_entries, xpath_map)
    merge_nullflavors(mfds_entries, nullflavor_usage)
    mfds_vocab = merge_vocabulary(mfds_entries, vocabulary_usage)
    path = write_dictionary("mfds-regional.json", "MFDS", MFDS_SOURCE, mfds_entries)
    print(
        f"wrote {path} ({len(mfds_entries)} KR extension elements,"
        f" {xpath_annotated} with xpath, {mfds_vocab} with a vocabulary)"
    )

    fda_entries = parse_fda_csv(fda_text)
    fda_sev_annotated = merge_fda_severity(fda_entries, fda_severity)
    path = write_dictionary("fda-regional.json", "FDA", FDA_SOURCE, fda_entries)
    elements = sum(1 for entry in fda_entries if entry["kind"] == "element")
    print(
        f"wrote {path} ({elements} elements, {len(fda_entries) - elements} groups,"
        f" {fda_sev_annotated} with FDA severity)"
    )

    for name, authority, source, rules in (
        ("ich.json", "ICH", ICH_SOURCE, extract_ich_rules(ich_text)),
        ("mfds.json", "MFDS", MFDS_SOURCE, extract_mfds_rules(mfds_sheet1)),
        ("fda.json", "FDA", FDA_SOURCE, extract_fda_rules(fda_text)),
    ):
        path = write_rules(name, authority, source, rules)
        print(f"wrote {path} ({len(rules)} rules)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
